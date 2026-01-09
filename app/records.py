from aiogram import F, Router
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, Message, CallbackQuery, FSInputFile, ReplyKeyboardRemove, KeyboardButton, ReplyKeyboardMarkup
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext
from app.states import Register
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
import uuid
import time
from aiogram.types import InputFile
import aiosqlite 
from app.data_shops import shops
import pandas as pd
import os  # Для работы с файлами и папками
import logging
from dotenv import load_dotenv
import json
from app.database import search_data, get_today_history
from googleapiclient.discovery import build  # Для Drive API
import io  # Для работы с BytesIO
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import app.keyboards as kb
import asyncio



router_records = Router()
load_dotenv('token.env')  # Загружаем переменные окружения из .env файла
logger = logging.getLogger(__name__)

# Путь к файлу, где будут храниться данные
FILE_PATH = 'json/machines_data.json'
FILE_PATH_ACCESS = 'json/access_user.json'
DRIVE_FILES_PATH = 'json/drive_files.json'
spreadsheet_id = os.getenv('GOOGLE_SHEET_KEY')
credentials_path = os.getenv('GOOGLE_CREDENTIALS_PATH')
# Папка для временных файлов
TEMP_DIR = 'temp files'
TEMP_FOLDER_ID = '1ihS9eD7QHZa0xsru_VKq_YKuEnN3T3iA'

# Функция для загрузки данных из JSON файла


def load_access_data():
    """Загружает данные пользователей из JSON-файла или создает структуру, если файл пуст/не существует."""
    try:
        with open(FILE_PATH_ACCESS, "r", encoding="utf-8") as file:
            return json.load(file)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        logger.warning(f"Файл доступа не найден или поврежден: {e}")
        return {
            "main_admins": [],
            "admins": [],
            "users": []
        }


# Функция сохранения истории файлов в JSON


def save_drive_files(files_list):
    """Сохраняет список файлов в JSON."""
    with open(DRIVE_FILES_PATH, "w", encoding="utf-8") as file:
        json.dump(files_list, file, ensure_ascii=False, indent=4)

# функция определения уровня доступа


def get_user_role(user_id, data):
    if user_id in data['main_admins']:
        return "👑 Главный администратор!"
    elif user_id in data['admins']:
        return "🛠 Администратор!"
    elif user_id in data['users']:
        return "👥 Пользователь"
    return None



# Inline кнопка Главное меню
inline_main_menu = InlineKeyboardMarkup(
    inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Главное меню", callback_data="main_menu")]
    ]
)


async def load_db_data():
    """Загружает все записи из БД (асинхронно)."""
    return await search_data("")

async def run_search(phrase):
    results = await search_data(phrase)
    # Добавляем индекс строки, если нужно (для редактирования)
    for idx, row in enumerate(results):
        row["__row"] = idx + 1  # Нумерация с 1
    return results

# Регистрируем шрифт DejaVu Sans (предполагаем, что файл DejaVuSans.ttf в корне проекта)
pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))

# Создаём стиль для параграфов с поддержкой кириллицы (для ячеек таблицы)
styles = getSampleStyleSheet()
normal_style = ParagraphStyle(
    'Normal',
    parent=styles['Normal'],
    fontName='DejaVuSans',  # Используем зарегистрированный шрифт
    fontSize=7,  # Уменьшаем шрифт для компактности
    leading=8,  # Межстрочный интервал
)

# Создаём стиль для заголовка (центрированный, больший шрифт, с отступами)
title_style = ParagraphStyle(
    'Title',
    parent=styles['Title'],  # Или 'Normal', если 'Title' не определён
    # Можно заменить на 'DejaVuSans-Bold' если есть файл DejaVuSans-Bold.ttf
    fontName='DejaVuSans',
    fontSize=12,  # Увеличенный шрифт для заголовка
    alignment=1,  # 1 = центр (0 = лево, 2 = право)
    spaceAfter=20,  # Отступ после заголовка (в pt, для разделения от таблицы)
    spaceBefore=0,  # Отступ перед заголовком (0 = без отступа сверху)
    textColor=colors.red,  # Цвет текста
)

# # Функция создания PDF файла


def create_pdf_file(results, filename):
    """Создает PDF файл с результатами поиска и возвращает путь к нему."""
    if not results:
        return None

    # Создаём папку, если её нет
    os.makedirs(TEMP_DIR, exist_ok=True)

    # Полный путь к файлу
    file_path = os.path.join(TEMP_DIR, filename)

    # Создаём DataFrame из результатов
    df = pd.DataFrame(results)


    column_rename = {
        'date': 'Дата',
        'workers': 'Исполнители работ',
        'work_description': 'Описание проблемы',
        'work_solution': 'Решение',
        'fault_status': 'Статус неисправности',
        'start_time': 'Дата начала',
        'end_time': 'Дата окончания',
        'duration': 'Затраченное время',
        'shift': 'Цех',
        'machine': 'Станок',
        'inventory_number': 'Инвентарный номер'
    }
    # Удаляем столбец id, если он есть (не нужен в выводе)
    if 'id' in df.columns:
        df = df.drop(columns=['id'])
    df = df.rename(columns=column_rename)

    # Создаём PDF документ с ландшафтной ориентацией для большего пространства
    doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
    elements = []

    # Заголовок
    search_phrase = filename.split('_')[2].replace('_', ' ') if len(filename.split('_')) > 2 else 'Запрос'
    title = Paragraph(f"Результаты поиска: '{search_phrase}'", title_style)
    elements.append(title)

    # Преобразуем DataFrame в список списков с Paragraph для каждой ячейки
    data = []
    for row in [df.columns.tolist()] + df.values.tolist():  # Заголовки + данные
        data_row = []
        for cell in row:
            cell_text = str(cell) if cell is not None else ""
            data_row.append(Paragraph(cell_text, normal_style))
        data.append(data_row)

    # Создаём таблицу с фиксированной шириной столбцов
    num_cols = len(df.columns)
    col_widths = [60, 50, 180, 180, 80, 40, 40, 40, 30, 40, 40]  # Расширенные настройки ширины
    
    # Автоподбор ширины для очень длинных таблиц
    total_width = sum(col_widths)
    page_width = 770  # Ширина страницы A4 в ландшафтном режиме (примерно)
    table = Table(data, colWidths=col_widths)

    # Стиль таблицы
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        # Автоматический перенос текста в ячейках
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ])
    table.setStyle(style)

    elements.append(table)

    # Генерируем PDF
    doc.build(elements)

    return file_path



def get_oauth_drive_service():
    """Возвращает аутентифицированный сервис для работы с Google Drive API через OAuth"""
    SCOPES = ['https://www.googleapis.com/auth/drive']
    
    creds = None
    token_path = 'token.json'
    
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)
    
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'json/OAUTH.json', SCOPES)
            creds = flow.run_local_server(port=0)
        
        with open(token_path, 'w') as token:
            token.write(creds.to_json())
    
    return build('drive', 'v3', credentials=creds)


# Функция создания Google Таблицы и сохранения копии в папку TEMP
def create_google_sheet(results, phrase, user_id):
    """Создает новую Google Таблицу с результатами поиска и сохраняет копию в папку TEMP"""
    if not results:
        logger.warning("Нет данных для создания таблицы.")
        return None

    try:
        # Аутентификация с помощью OAuth
        client = connect_to_google_sheets()
        
        # Получаем credentials напрямую из функции connect_to_google_sheets
        # Для этого нужно немного изменить connect_to_google_sheets, чтобы она возвращала и creds
        # Либо получаем creds здесь заново
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
                 'https://www.googleapis.com/auth/drive']
        
        creds = None
        token_path = 'token.json'
        
        if os.path.exists(token_path):
            creds = Credentials.from_authorized_user_file(token_path, SCOPES)
        
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'json/OAUTH.json', SCOPES)
                creds = flow.run_local_server(port=0)
            
            with open(token_path, 'w') as token:
                token.write(creds.to_json())

        # Создаем низкоуровневый сервис для Sheets API
        sheets_service = build('sheets', 'v4', credentials=creds)

        # Создаем имя для таблицы
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        sheet_name = f"Результаты_поиска_{phrase}_{timestamp}"

        # Создаем новую таблицу через gspread
        new_spreadsheet = client.create(sheet_name)
        spreadsheet_id = new_spreadsheet.id
        logger.info(f"Таблица создана: {sheet_name} (ID: {spreadsheet_id})")

        # Записываем данные в таблицу
        worksheet = new_spreadsheet.sheet1
        df = pd.DataFrame(results)
        
        # # Подготавливаем данные: заголовки и строки
        data_to_update = [df.columns.values.tolist()] + df.values.tolist()
        worksheet.update(data_to_update)
        logger.info("Данные успешно записаны в таблицу.")
        sheet_id = int(worksheet.id)  # gspread возвращает реальный sheetId

        num_rows = len(data_to_update)
        num_cols = len(data_to_update[0]) if data_to_update else 0

        full_range = {
            "sheetId": sheet_id,
            "startRowIndex": 0,
            "endRowIndex": num_rows,
            "startColumnIndex": 0,
            "endColumnIndex": num_cols
        }

        header_range = {
            "sheetId": sheet_id,
            "startRowIndex": 0,
            "endRowIndex": 1,
            "startColumnIndex": 0,
            "endColumnIndex": num_cols
        }

        # ЗАМЕНА: Убираем autoResizeDimensions и добавляем индивидуальные настройки ширины
        column_width_requests = [
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 0,  # Колонка 0: Дата
                        "endIndex": 1
                    },
                    "properties": {"pixelSize": 120},
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 1,  # Колонка 1: Исполнители
                        "endIndex": 2
                    },
                    "properties": {"pixelSize": 150},
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 2,  # Колонка 2: Описание проблемы
                        "endIndex": 3
                    },
                    "properties": {"pixelSize": 400},  # Широкая для текста
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 3,  # Колонка 3: Решение
                        "endIndex": 4
                    },
                    "properties": {"pixelSize": 400},  # Широкая для текста
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 4,  # Колонка 4: Статус
                        "endIndex": 5
                    },
                    "properties": {"pixelSize": 150},
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 5,  # Колонка 5: Начало работ
                        "endIndex": 6
                    },
                    "properties": {"pixelSize": 150},
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 6,  # Колонка 6: Окончание работ
                        "endIndex": 7
                    },
                    "properties": {"pixelSize": 150},
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 7,  # Колонка 7: Затраченное время
                        "endIndex": 8
                    },
                    "properties": {"pixelSize": 120},
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 8,  # Колонка 8: Цех
                        "endIndex": 9
                    },
                    "properties": {"pixelSize": 100},
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 9,  # Колонка 9: Станок
                        "endIndex": 10
                    },
                    "properties": {"pixelSize": 120},
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 10,  # Колонка 10: Инвентарный номер
                        "endIndex": 11
                    },
                    "properties": {"pixelSize": 180},
                    "fields": "pixelSize"
                }
            }
        ]

        requests = [
            {
                "setBasicFilter": {
                    "filter": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": num_rows,
                            "startColumnIndex": 0,
                            "endColumnIndex": num_cols
                        }
                    }
                }
            },
            {
                "repeatCell": {
                    "range": full_range,
                    "cell": {
                        "userEnteredFormat": {
                            "wrapStrategy": "WRAP",
                            "horizontalAlignment": "CENTER",
                            "verticalAlignment": "MIDDLE"
                        }
                    },
                    "fields": "userEnteredFormat(wrapStrategy, horizontalAlignment, verticalAlignment)"
                }
            },
            {
                "repeatCell": {
                    "range": header_range,
                    "cell": {
                        "userEnteredFormat": {
                            "textFormat": {"bold": True},
                            "backgroundColor": {"red": 0.9, "green": 0.9, "blue": 0.9}
                        }
                    },
                    "fields": "userEnteredFormat(textFormat, backgroundColor)"
                }
            },
            # ЗАМЕНА: добавляем наши индивидуальные настройки ширины
            *column_width_requests,
            {
                "addProtectedRange": {
                    "protectedRange": {
                        "range": header_range,
                        "description": "Защита строки заголовков",
                        "warningOnly": False,
                        "requestingUserCanEdit": False,
                        "editors": {
                            "users": [], 
                            "groups": [],
                            "domainUsersCanEdit": False
                        }
                    }
                }
            }
        ]

        # Выполняем batchUpdate
        sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests}
        ).execute()
        logger.info("Форматирование и защита заголовков успешно применены.")

        # Перемещаем файл в папку TEMP
        if TEMP_FOLDER_ID:
            try:
                drive_service = build('drive', 'v3', credentials=creds)
                
                # Перемещаем файл из корня в указанную папку
                drive_service.files().update(
                   fileId=spreadsheet_id,
                    addParents=TEMP_FOLDER_ID,
                    removeParents='root',
                    fields='id, parents'
                ).execute()
                logger.info(f"Файл успешно перемещен в папку TEMP: {TEMP_FOLDER_ID}")
                
            except Exception as move_error:
                logger.error(f"Не удалось переместить файл в папку TEMP. Ошибка: {move_error}")
        else:
            logger.warning("TEMP_FOLDER_ID не указан. Файл останется в корневой папке.")

        # Делаем таблицу доступной всем
        new_spreadsheet.share(None, perm_type='anyone', role='writer')
        logger.info("Таблица стала доступной для чтения по ссылке.")

        # Формируем ссылку вручную
        manual_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
        # return manual_url
        return {
                "url": manual_url,
                "copy_sheet_id": spreadsheet_id,
                "row_map": [row["__row"] for row in results]}  # список исходных строк
    

    except Exception as e:
        logger.error(f"Критическая ошибка при создании Google Таблицы: {e}")
        return None




def cleanup_old_files():
    """Удаляет файлы из TEMP_DIR старше 24 часов."""
    if not os.path.exists(TEMP_DIR):
        return

    now = time.time()
    for filename in os.listdir(TEMP_DIR):
        # Удаляем и .xlsx (история Google Таблиц?) и .pdf (результаты поиска)
        if filename.endswith('.pdf'):
            file_path = os.path.join(TEMP_DIR, filename)
            file_time = os.path.getctime(file_path)
            if now - file_time > 86400:
                os.remove(file_path)
                logger.info(f'Файл {filename} удален.')


# Обработчик кнопки "🔍 Поиск записи" — запрашивает фразу и переходит в состояние


@router_records.message(F.text == '🔍 Поиск записи')
async def start_search(message: Message, state: FSMContext):
    data = load_access_data()  # Загружаем данные о пользователях
    user_id = message.from_user.id  # Получаем ID пользователя
    role = get_user_role(user_id, data)
    if role is None:
        await message.answer("Доступ запрещён.")
        return

    logger.info(f"Пользователь {user_id} ({role}) начал поиск записи.")
    await message.answer("Введите слово или фразу для поиска по базе (не может быть пустым):", reply_markup=ReplyKeyboardRemove())
    # Используем ваше существующее состояние
    await state.set_state(Register.search_record)


@router_records.message(StateFilter(Register.search_record))
async def process_search_phrase(message: Message, state: FSMContext):
    phrase = message.text.strip()
    if not phrase:
        return await message.answer(
            "Фраза не может быть пустой. Введите заново:",
            reply_markup=inline_main_menu
        )

    # Отправляем первое сообщение о прогрессе
    progress_msg = await message.answer("🔍 Идёт поиск, пожалуйста подождите...")

    try:
        # Этап 1 — поиск (используем нашу функцию search_data вместо run_search)
        results = await search_data(phrase)
        await asyncio.sleep(0.5)
        await progress_msg.edit_text("⏳ Обработка результатов...")

        if not results:
            await progress_msg.delete()
            await message.answer(
                f"По запросу '{phrase}' ничего не найдено.\nВведите новую фразу:",
                reply_markup=inline_main_menu
            )
            return

        # Этап 2 — создание PDF
        await asyncio.sleep(0.5)
        await progress_msg.edit_text("📄 Формирую файл с результатами...")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Результат_{message.from_user.id}_{phrase}_{timestamp}.pdf"  # Изменил на .pdf, так как создаём PDF
        file_path = create_pdf_file(results, filename)

        # Этап 3 — финал
        await asyncio.sleep(0.5)
        await progress_msg.edit_text("🧾 Подготавливаю отправку результата...")

        # Удаляем индикатор
        await progress_msg.delete()

        # Отправляем PDF
        await message.answer_document(
            document=FSInputFile(file_path),
            caption=f"По запросу '{phrase}' найдено {len(results)} результатов.",
            reply_markup=inline_main_menu
        )

        await state.clear()

    except Exception as e:
        logger.error(f"Ошибка при поиске: {e}")  # Логируем для отладки
        await progress_msg.edit_text("❌ Ошибка при обработке запроса.")
        await state.clear()
        await message.answer(
            f"Ошибка: {str(e)}. Попробуйте позже.",
            reply_markup=inline_main_menu
        )

@router_records.callback_query(lambda c: c.data == "main_menu")
async def go_to_main_menu(callback: CallbackQuery):
    try:
        # Удаляем сообщение с PDF и кнопкой
        await callback.message.delete()
    except Exception as e:
        # Иногда сообщение может быть уже удалено, тогда просто логируем
        logger.warning(f"Не удалось удалить сообщение: {e}")

    # Отправляем главное меню
    await callback.message.answer(
        "Главное меню:",
        reply_markup=kb.main  # твой ReplyKeyboardMarkup
    )

    # Заканчиваем callback
    await callback.answer()


def create_local_excel(results, phrase, user_id):
    """Создает локальный Excel-файл с результатами поиска, форматирует его и сохраняет в temp_files."""
    if not results:
        logger.warning("Нет данных для создания файла.")
        return None

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты поиска"

        df = pd.DataFrame(results)
        data_to_update = [df.columns.values.tolist()] + df.values.tolist()
        for row_num, row_data in enumerate(data_to_update, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value

        num_rows = len(data_to_update)
        num_cols = len(data_to_update[0]) if data_to_update else 0
        full_range = f"A1:{get_column_letter(num_cols)}{num_rows}"
        header_range = f"A1:{get_column_letter(num_cols)}1"

        # Ширина колонок (в пикселях, преобразованная в символы)
        column_widths = [120, 150, 400, 400, 150, 150, 150, 120, 100, 120, 180]
        for i, width_pixels in enumerate(column_widths[:num_cols]):
            width_chars = width_pixels / 8  # Примерное преобразование
            ws.column_dimensions[get_column_letter(i+1)].width = width_chars

        ws.auto_filter.ref = full_range
        for row in ws[full_range]:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        ws.protection.sheet = True
        for row in ws.iter_rows(min_row=2, max_row=num_rows):
            for cell in row:
                cell.protection = Protection(locked=False)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        safe_phrase = "".join(c for c in phrase if c.isalnum() or c in "_- ").replace(" ", "_")[:50]
        file_name = f"{user_id}_{safe_phrase}_{timestamp}_{uuid.uuid4().hex[:8]}.xlsx"
        file_path = os.path.join(TEMP_DIR, file_name)

        wb.save(file_path)
        logger.info(f"Файл создан: {file_path}")

        row_map = [row.get("id") for row in results]  # Используем 'id' из SQLite
        return {
            "file_path": file_path,
            "row_map": row_map,
            "original_data": results  # Сохраняем оригинал для сравнения
        }

    except Exception as e:
        logger.error(f"Ошибка при создании Excel-файла: {e}")
        return None


    
    
async def update_record_in_db(record_id, updated_data):
    """
    Асинхронно обновляет запись в SQLite по id.
    
    :param record_id: int — ID записи для обновления.
    :param updated_data: dict — Словарь с полями для обновления.
    """
    try:
        conn = await aiosqlite.connect('bot_data.db')  # Путь к вашей БД
        cursor = await conn.cursor()
        
        # Формируем SET-часть запроса динамически
        set_clause = ', '.join([f"{k} = ?" for k in updated_data.keys()])
        values = list(updated_data.values()) + [record_id]  # Добавляем ID
        
        # Выполняем UPDATE
        query = f"UPDATE tasks SET {set_clause} WHERE id = ?"  
        await cursor.execute(query, values)
        
        # Сохраняем изменения
        await conn.commit()
        
        # Логируем успех
        logger.info(f"Запись с ID {record_id} обновлена: {updated_data}")
        
    except aiosqlite.Error as e:
        logger.error(f"Ошибка при обновлении записи ID {record_id}: {e}")
        raise  # Перебрасываем для обработки
    finally:
        if conn:
            await conn.close()
    

@router_records.message(F.text == '✏️ Изменить запись')
async def start_edit(message: Message, state: FSMContext):
    data = load_access_data()
    user_id = message.from_user.id
    role = get_user_role(user_id, data)
    if role is None:
        await message.answer("Доступ запрещён.")
        return

    logger.info(f"Пользователь {user_id} начал редактирование записи.")
    await message.answer("Введите слово или фразу для поиска:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(Register.waiting_for_search_phrase)


@router_records.message(StateFilter(Register.waiting_for_search_phrase))
async def process_search_phrase(message: Message, state: FSMContext):
    phrase = message.text.strip()

    if not phrase:
        return await message.answer("Фраза не может быть пустой. Попробуйте ещё раз:")

    # Отправляем первое сообщение о прогрессе
    progress_msg = await message.answer("🔍 Идёт поиск, пожалуйста подождите...")

    try:
        results = await run_search(phrase)
        await asyncio.sleep(0.5)
        await progress_msg.edit_text("⏳ Обработка результатов...")
        await asyncio.sleep(0.5)  # Пауза после обработки

        if not results:
            await progress_msg.delete()
            return await message.answer(
                f"🔍 По запросу '<code>{phrase}</code>' ничего не найдено.\n\n"
                f"• Попробуйте ввести другой запрос\n"
                f"• Или вернитесь в главное меню",
                reply_markup=inline_main_menu,
                parse_mode="HTML"
            )

        # Сохраняем результаты и начинаем показ первой записи
        await state.update_data(search_results=results, current_index=0, search_phrase=phrase)
        await progress_msg.edit_text("📄 Подготовка к показу результатов...")
        await asyncio.sleep(0.3)  # Небольшая пауза перед открытием
        await progress_msg.delete()
        await show_record(message, state)
        await state.set_state(Register.viewing_record)

    except Exception as e:
        logger.error(f"Ошибка при поиске: {e}")
        await progress_msg.edit_text("❌ Ошибка при обработке запроса.")
        await state.clear()
        await message.answer(
            f"Ошибка: {str(e)}. Попробуйте позже.",
            reply_markup=inline_main_menu
        )


async def show_record(message: Message, state: FSMContext):
    data = await state.get_data()
    results = data["search_results"]
    index = data["current_index"]
    record = results[index].copy()
    

    total = len(results)
    msg_text = (
        f"🚀 <b>ЗАЯВКА</b> <code>#{record['id']}</code>\n"
        f"📱 <b>СТРАНИЦА:</b> <code>{index + 1}/{total}</code>\n"
        f"{'•' * 30}\n"
        f"📅 <b>Дата:</b> {record['date']}\n"
        f"📌 <b>Исполнители работ:</b> {record['workers']}\n"
        f"📝 <b>Описание проблемы:</b> {record['work_description']}\n"
        f"📝 <b>Решение:</b> {record['work_solution']}\n"
        f"📝 <b>Статус неисправности:</b> {record['fault_status']}\n"
        f"📅 <b>Дата начала:</b> {record['start_time']}\n"
        f"📅 <b>Дата окончания:</b> {record['end_time']}\n"
        f"⏳ <b>Затраченное время:</b> {record['duration']}\n"
        f"🏭 <b>Цех:</b> {record['shift']}\n"
        f"🔧 <b>Станок:</b> {record['machine']}\n"
        f"🔢 <b>Инвентарный номер:</b> {record['inventory_number']}"
    )

    keyboard = build_navigation_buttons(index, total)
    if isinstance(message, CallbackQuery):
        await message.message.edit_text(msg_text, reply_markup=keyboard, parse_mode="HTML")
    else:
        await message.answer(msg_text, reply_markup=keyboard, parse_mode="HTML")
        

    


def build_navigation_buttons(current_index, total):
    buttons = []

    # Кнопки редактирования
    edit_buttons = [
        [InlineKeyboardButton(text="🔧 Изм. проблему", callback_data="edit_problem"),
         InlineKeyboardButton(text="🛠 Изм. решение", callback_data="edit_solution")],
        [InlineKeyboardButton(text="📊 Изм. статус", callback_data="edit_status"),
         InlineKeyboardButton(text="👷 Изм. исполнителей", callback_data="edit_workers")]
    ]
    buttons.extend(edit_buttons)

    nav_buttons = []
    if current_index > 0:
        nav_buttons.append(InlineKeyboardButton(text="⬅️ Предыдущая", callback_data="prev_record"))
    if current_index < total - 1:
        nav_buttons.append(InlineKeyboardButton(text="➡️ Следующая", callback_data="next_record"))

    if nav_buttons:
        buttons.append(nav_buttons)
    
    buttons.append([InlineKeyboardButton(text="🔙 Главное меню", callback_data="main_menu")])

    return InlineKeyboardMarkup(inline_keyboard=buttons)


# Обработка перехода между записями
@router_records.callback_query(F.data.in_({"prev_record", "next_record"}))
async def navigate_records(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    index = data["current_index"]
    total = len(data["search_results"])

    if callback.data == "prev_record" and index > 0:
        await state.update_data(current_index=index - 1)
    elif callback.data == "next_record" and index < total - 1:
        await state.update_data(current_index=index + 1)
    else:
        await callback.answer()
        return

    await show_record(callback, state)
    await callback.answer()


# Обработка начала редактирования поля
@router_records.callback_query(F.data.startswith("edit_"))
async def start_field_edit(callback: CallbackQuery, state: FSMContext):
    field_map = {
        "edit_problem": ("work_description", "Введите новое описание проблемы:"),
        "edit_solution": ("work_solution", "Введите новое решение:"),
        "edit_status": ("fault_status", "Введите новый статус:"),
        "edit_workers": ("workers", "Введите новых исполнителей работ:")
    }

    field_key, prompt = field_map[callback.data]
    data = await state.get_data()
    current_index = data["current_index"]
    records = data["search_results"]
    old_value = records[current_index][field_key]

    await state.update_data(editing_field=field_key, old_value=old_value)
    
    # Создаем кнопку для копирования старого текста
    copy_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 Скопировать текущий текст", callback_data="copy_old_text")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_edit_field")]
    ])

    await callback.message.answer(
        "✅ Текст готов к копированию!\n\n"
        "🔹 <b>Что делать дальше:</b>\n"
        "• Нажмите кнопку 'Скопировать' ниже\n"
        "• Скопируйте текст нажатием на текст\n"
        "• Вставьте его в поле ввода ниже ⬇️\n"
        "• Внесите необходимые изменения\n"
        "• Отправьте сообщение и подтвердите для сохранения\n\n"
        "<i>Или просто введите новую информацию вручную</i>",
        reply_markup=copy_kb,
        parse_mode="HTML"
    )
    await callback.answer()


# Обработка отмены редактирования поля
@router_records.callback_query(F.data == "cancel_edit_field")
async def cancel_field_edit(callback: CallbackQuery, state: FSMContext):
    # Очищаем данные редактирования
    await state.update_data(editing_field=None, old_value=None, new_value=None)
    
    # Возвращаемся к просмотру записи
    await state.set_state(Register.viewing_record)
    await show_record(callback, state)
    await callback.answer()


# Обработка копирования старого текста
@router_records.callback_query(F.data == "copy_old_text")
async def copy_old_text(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    old_value = data["old_value"]
    
    # Отправляем старый текст как сообщение, которое пользователь может скопировать
    await callback.message.edit_text(
        f"\n\n<code>{old_value}</code>\n\n",
        parse_mode="HTML"
    )
    
    # Убираем кнопки и ждем ввода нового значения
    await state.set_state(Register.editing_field)
    await callback.answer()


# Обработка нового значения поля
@router_records.message(StateFilter(Register.editing_field))
async def save_edited_field(message: Message, state: FSMContext):
    new_value = message.text.strip()
    if not new_value:
        return await message.answer("Значение не может быть пустым. Попробуйте ещё раз:")

    data = await state.get_data()
    field_to_update = data["editing_field"]
    old_value = data["old_value"]

    # Сохраняем новое значение временно
    await state.update_data(new_value=new_value)

    # Создаём кнопки подтверждения
    confirm_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Сохранить", callback_data="confirm_save")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_save")]
    ])

    await message.answer(
        f"Вы хотите изменить поле на:\n\n<b>{new_value}</b>\n\nВыберите действие:",
        reply_markup=confirm_kb,
        parse_mode="HTML"
    )
    await state.set_state(Register.confirming_edit)
    
# Подтверждение сохранения изменений
@router_records.callback_query(F.data == "confirm_save", StateFilter(Register.confirming_edit))
async def confirm_save(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    field_to_update = data["editing_field"]
    new_value = data["new_value"]
    current_index = data["current_index"]
    records = data["search_results"]
    record = records[current_index]

    # Обновляем значение в памяти
    record[field_to_update] = new_value
    records[current_index] = record
    await state.update_data(search_results=records)

    # Сохраняем в БД
    try:
        await update_record_in_db(record["id"], {field_to_update: new_value})
        await callback.message.edit_text("✅ Поле успешно обновлено!", reply_markup=None)
    except Exception as e:
        logger.error(f"Ошибка при обновлении записи: {e}")
        await callback.message.edit_text("❌ Произошла ошибка при сохранении.", reply_markup=None)

    await state.set_state(Register.viewing_record)
    await show_record(callback, state)
    await callback.answer()


# Отмена сохранения изменений
@router_records.callback_query(F.data == "cancel_save", StateFilter(Register.confirming_edit))
async def cancel_save(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("↩️ Изменения отменены.", reply_markup=None)
    await state.set_state(Register.viewing_record)
    await show_record(callback, state)
    await callback.answer()



# @router_records.message(F.text == '✏️ Изменить запись')
# async def start_edit(message: Message, state: FSMContext):
#     data = load_access_data()
#     user_id = message.from_user.id
#     role = get_user_role(user_id, data)
#     if role is None:
#         await message.answer("Доступ запрещён.")
#         return

#     logger.info(f"Пользователь {user_id} начал редактирование записи.")
#     await message.answer("Введите слово или фразу для поиска по базе (не может быть пустым):", reply_markup=ReplyKeyboardRemove())
#     await state.set_state(Register.edit_record)

# # Обработчик фразы поиска
# @router_records.message(StateFilter(Register.edit_record))
# async def process_edit_phrase(message: Message, state: FSMContext):
#     phrase = message.text.strip()

#     if not phrase:
#         return await message.answer(
#             "Фраза не может быть пустой. Попробуйте ещё раз:",
#             reply_markup=inline_main_menu
#         )

#     progress = await message.answer("🔍 Идёт поиск, пожалуйста подождите...")

#     try:
#         results = await run_search(phrase)

#         if not results:
#             await progress.delete()
#             return await message.answer(
#                 f"По запросу '{phrase}' ничего не найдено.\n"
#                 f"Введите новую фразу или нажмите кнопку ниже:",
#                 reply_markup=inline_main_menu
#             )

#         await progress.edit_text("⏳ Создание файла, пожалуйста подождите...")

#         # Создание локального Excel
#         file_info = create_local_excel(results, phrase, message.from_user.id)

#         if not file_info:
#             await progress.delete()
#             await state.clear()
#             return await message.answer(
#                 "Ошибка: Не удалось создать файл. Попробуйте позже.",
#                 reply_markup=inline_main_menu
#             )

#         # Сохраняем данные в состоянии
#         await state.update_data(
#             file_path=file_info["file_path"],
#             row_map=file_info["row_map"],
#             original_data=file_info["original_data"]
#         )

#         # Отправляем файл пользователю
#         await message.answer_document(
#         document=FSInputFile(file_info["file_path"], filename="results.xlsx"),  # Используем FSInputFile с путём и именем файла
#         caption=f"Найдено {len(results)} строк по запросу '{phrase}'.\n"
#             "Скачайте файл, внесите изменения и загрузите обратно."
#     )

#         # Клавиатура для загрузки/отмены
#         keyboard = InlineKeyboardMarkup(
#             inline_keyboard=[
#                 [InlineKeyboardButton(text="📤 Загрузить изменения", callback_data="upload_edit")],
#                 [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_edit")]
#             ]
#         )

#         await progress.delete()

#         result_msg = await message.answer(
#             "Файл отправлен. После редактирования нажмите «Загрузить изменения» и пришлите файл.",
#             reply_markup=keyboard
#         )

#         await state.update_data(result_message_id=result_msg.message_id)

#     except Exception as e:
#         logger.error(f"Ошибка в process_edit_phrase: {e}", exc_info=True)
#         await progress.delete()
#         await state.clear()
#         await message.answer(
#             f"Произошла ошибка: {str(e)}. Попробуйте позже.",
#             reply_markup=inline_main_menu
#         )

# # Обработчик выбора записи

# @router_records.message(StateFilter(Register.edit_record), F.document)
# async def handle_uploaded_file(message: Message, state: FSMContext):
#     data = await state.get_data()
#     file_path = data.get("file_path")
#     row_map = data.get("row_map")
#     original_data = data.get("original_data")

#     if not file_path or not row_map:
#         return await message.answer("Сначала выполните поиск и создайте файл.", reply_markup=inline_main_menu)

#     # Проверяем тип файла
#     if not message.document.file_name.endswith('.xlsx'):
#         return await message.answer("Пожалуйста, загрузите файл в формате .xlsx.")

#     try:
#         # Получаем file_id
#         file_id = message.document.file_id
        
#         # Получаем информацию о файле
#         file_info = await message.bot.get_file(file_id)
        
#         # Скачиваем файл
#         downloaded_file = await message.bot.download_file(file_info.file_path)
        
#         # Сохраняем временно для чтения (downloaded_file — это BytesIO)
#         temp_upload_path = os.path.join(TEMP_DIR, f"uploaded_{uuid.uuid4().hex}.xlsx")
#         with open(temp_upload_path, 'wb') as f:
#             f.write(downloaded_file.read())

#         # Читаем данные из загруженного файла
#         df_uploaded = pd.read_excel(temp_upload_path, header=0)
#         uploaded_rows = df_uploaded.to_dict('bot_data')

#         # Удаляем временный загруженный файл
#         os.remove(temp_upload_path)

#         # Проверяем количество строк
#         if len(uploaded_rows) != len(original_data):
#             return await message.answer("Количество строк в файле не совпадает с оригиналом. Проверьте файл.")

#         # Сравниваем и обновляем
#         changes_made = False
#         for i, (orig, uploaded) in enumerate(zip(original_data, uploaded_rows)):
#             # Убираем 'id' из сравнения (он не редактируется)
#             orig_clean = {k: v for k, v in orig.items() if k != 'id'}
#             uploaded_clean = {k: v for k, v in uploaded.items() if k in orig_clean}
            
#             if orig_clean != uploaded_clean:
#                 # Обновляем запись в SQLite
#                 await update_record_in_db(row_map[i], uploaded_clean)
#                 changes_made = True

#         # Удаляем оригинальный файл
#         if os.path.exists(file_path):
#             os.remove(file_path)

#         if changes_made:
#             await message.answer("✅ Изменения успешно сохранены!", reply_markup=inline_main_menu)
#         else:
#             await message.answer("🗑️ Изменений не найдено.", reply_markup=inline_main_menu)

#         await state.clear()

#     except Exception as e:
#         logger.error(f"Ошибка при обработке файла: {e}", exc_info=True)
#         await message.answer("❌ Ошибка при обработке файла. Попробуйте снова.", reply_markup=inline_main_menu)


# # Callback для отмены (удаляем файл, очищаем состояние)
# @router_records.callback_query(F.data == "cancel_edit")
# async def cancel_edit(callback: CallbackQuery, state: FSMContext):
#     data = await state.get_data()
#     file_path = data.get("file_path")

#     try:
#         await callback.message.edit_reply_markup(reply_markup=None)
#         if file_path and os.path.exists(file_path):
#             os.remove(file_path)
#         await callback.message.answer("❌ Редактирование отменено. Файл удалён.", reply_markup=inline_main_menu)
#     except Exception as e:
#         logger.error(f"Ошибка при отмене: {e}", exc_info=True)
#         await callback.message.answer("⚠️ Ошибка при отмене.", reply_markup=inline_main_menu)

#     await state.clear()
#     await callback.answer()

# # Заглушка для upload_edit (просто инструкция, файл обрабатывается в handle_uploaded_file)
# @router_records.callback_query(F.data == "upload_edit")
# async def upload_edit(callback: CallbackQuery):
#     await callback.message.answer("Пришлите изменённый файл .xlsx в ответ на это сообщение.")
#     await callback.answer()